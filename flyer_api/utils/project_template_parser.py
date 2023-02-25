from openpyxl import Workbook, load_workbook
from products_catalog_connector.catalog import Catalog
from django.conf import settings

# solo codice e prezzo obbligatori
class ProjectTemplateParser:
    row_index_start = 4

    def parse(project: 'flyer_api.models.Project'):
        # if project is None:
        #     project = Project.objects.get(pk=1)
        file_path = project.project_template_file.path
        workbook = load_workbook(filename=file_path, read_only=True)
        sheet = workbook.get_sheet_by_name(workbook.sheetnames[0])
        # print(f"row count: {sheet.max_row}")
        project_page_added = []
        for index, row in enumerate(sheet.rows):
            if index >= ProjectTemplateParser.row_index_start:
                if row[0].value is None:
                    # print(project_page_added)
                    return
                if isinstance(row[0].value, int):
                    if (row[2].value is not None):
                        current_page_number = row[0].value
                        if current_page_number in project_page_added:
                            current_page = project.pages.filter(
                                number=current_page_number)[0]
                        else:
                            current_page = project.pages.create(
                                number=current_page_number)
                            project_page_added.append(current_page_number)

                        product_attibutes = {
                            "code": row[2].value,
                            "position": row[1].value,
                            "other_codes": row[3].value,
                            "additional_data": {
                                "template_description": row[4].value
                            },
                            "note": row[5].value,
                            "price": row[6].value,
                            "price_without_discount": row[7].value,
                            "discount_percentage": ((row[8].value * 100) if (row[8].value is not None) else None),
                            "pieces_number": row[9].value,
                            "loyalty": (row[10].value == "x"),
                            "focus": (row[11].value == "x"),
                            "stopper": (row[12].value == "x"),
                            "poster": (row[13].value == "x"),
                        }
                        # product_info = Catalog.get_product_by_code(code=product_attibutes["code"], seller_code=project.user.profile.seller_code)
                        catalog_product = Catalog.get_product_by_code(code=product_attibutes["code"], seller_code=project.user.profile.seller_code)
                        # product_attibutes["description"] = product_info["description"]
                        product_attibutes["description"] = catalog_product.description
                        product_attibutes["description_brand"] = catalog_product.description_brand
                        product_attibutes["description_type"] = catalog_product.description_type
                        product_attibutes["description_tastes"] = catalog_product.description_tastes
                        product_attibutes["description_weight"] = catalog_product.description_weight
                        product = current_page.products.create(**product_attibutes)
                        # for image in product_info["images"]:
                        for image in catalog_product.images:
                            # product.images.create(url=image)
                            product.images.create(url=image.url)
