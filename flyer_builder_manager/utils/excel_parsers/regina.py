import requests
from io import BytesIO

from django.core import files
from django.conf import settings

from openpyxl import Workbook, load_workbook
from products_catalog_connector.models import Block, Seller, Distribution


def parse(project, seller, excel_file_path):
    try:
        workbook = load_workbook(filename=excel_file_path, read_only=False)
        sheet = workbook.get_sheet_by_name(workbook.sheetnames[0])

        project_pages = {}
        current_page_number = 0
        parser_log = []

        for index, row in enumerate(sheet.rows):
            if index >= 2:
                if row[0].value == "codice template":
                    current_page_number = current_page_number + 1
                    project_pages[current_page_number] = {
                        "has_products": False,
                        "template_code": row[3].value,
                        "products": [],
                        "name": row[4].value
                    }
                elif isinstance(row[0].value, int) and (row[2].value is not None):
                    project_pages[current_page_number]["has_products"] = True
                    project_pages[current_page_number]["products"].append(
                        {
                            "code": row[2].value,
                            "position": row[1].value,
                            "tag_code": row[3].value,
                            "additional_data": {
                                "template_description": row[4].value
                            },
                            "note": row[5].value,
                            "price": row[6].value,
                            "flyer_block_code": row[7].value,
                            "price_without_discount": row[8].value,
                            "discount_percentage": ((row[9].value * 100) if (row[9].value is not None) else None),
                            "pieces_number": row[10].value,
                            "max_purchasable_pieces": row[11].value,
                            "points": row[12].value,
                            "stopper": (row[13].value == "x"),
                            "poster": (row[14].value == "x"),
                            "row_number": index
                        }
                    )

        for page_number, excel_page_info in project_pages.items():
            if excel_page_info["has_products"]:
                page = project.pages.create(
                    number=page_number, name=excel_page_info["name"], template=project.client.page_layout_templates.filter(code=excel_page_info["template_code"])[0])
                
                parser_log.append(f"<div class='font-weight-bold'>Pagina {page.number}</div>")

                page.create_cells()
                page_cells = page.cells.order_by('position')
                current_page_free_cells = {}
                for cell in page_cells:
                    current_page_free_cells[cell.position] = cell
                excess_products_count = 0
                for excel_product in excel_page_info["products"]:
                    if Distribution.objects.filter(code=excel_product["code"], seller=seller).exists():
                        distribution = Distribution.objects.get(
                            code=excel_product["code"], seller=seller)
                        catalog_product = distribution.product
                        current_cell = None
                        if excel_product["position"] in current_page_free_cells:
                            current_cell = current_page_free_cells.pop(
                                excel_product["position"])
                        else:
                            excess_products_count += 1
                        product = project.products.create(
                            cell=current_cell,
                            ref_id=catalog_product.pk,
                            code=excel_product["code"],
                            description=catalog_product.get_description(),
                            description_brand=catalog_product.get_description_brand(),
                            description_type=catalog_product.get_description_type(),
                            description_tastes=catalog_product.get_description_tastes(),
                            description_weight=catalog_product.get_description_weight(),
                            position=excel_product["position"],
                            tag_code=excel_product["tag_code"],
                            additional_data=excel_product["additional_data"],
                            note=excel_product["note"],
                            price=excel_product["price"],
                            price_without_discount=excel_product["price_without_discount"],
                            discount_percentage=excel_product["discount_percentage"],
                            pieces_number=excel_product["pieces_number"],
                            stopper=excel_product["stopper"],
                            poster=excel_product["poster"],
                            max_purchasable_pieces=excel_product["max_purchasable_pieces"],
                            points=excel_product["points"],
                        )
                        if Block.objects.filter(code=excel_product["flyer_block_code"], seller=seller).exists():
                            flyer_block = Block.objects.get(
                                code=excel_product["flyer_block_code"], seller=seller)
                            product.ref_block_id = flyer_block.pk
                            product.save()

                        if catalog_product.picture.filter(favorite=True).exists():
                            catalog_picture = catalog_product.picture.filter(favorite=True)[
                                0]
                            catalog_picture_url = settings.CATALOG_IMAGES_URL_PREFIX + \
                                catalog_picture.low_resolution_path()
                            catalog_picture_image_response = requests.get(
                                catalog_picture_url)
                            if catalog_picture_image_response.status_code == requests.codes.ok:
                                bytes_io = BytesIO()
                                bytes_io.write(
                                    catalog_picture_image_response.content)
                                file_name = catalog_picture_url.split("/")[-1]
                                product_image = product.images.create()
                                product_image.image.save(
                                    file_name, files.File(bytes_io))
                        else:
                            # no image found
                            parser_log.append(
                                f"<div class='text-danger'>- Riga numero {excel_product['row_number'] + 1}: immagine prodotto non trovata</div>")
                    else:
                        # no product found
                        parser_log.append(
                            f"<div class='text-danger'>- Riga numero {excel_product['row_number'] + 1}: prodotto con codice {excel_product['code']} non trovato</div>")
                if excess_products_count > 0:
                    parser_log.append(f"<div class='text-info'>- Prodotti eccedenti: {excess_products_count}</div>")    
                parser_log.append(f"<div>- Prodotti inseriti: {project.products.filter(cell__in=page.cells.all()).count()}</div>")
            else:
                if {k: v for (k, v) in project_pages.items() if k > page_number if v["has_products"]}:
                    page = project.pages.create(number=page_number, name=excel_page_info["name"], template=project.client.page_layout_templates.filter(
                        code=excel_page_info["template_code"])[0])
                    page.create_cells()
                    parser_log.append(f"<div class='font-weight-bold'>Pagina {page.number}</div>")
                    parser_log.append(f"<div>- Prodotti inseriti: 0</div>")

        if project.pages.count() == 0:
            page = project.pages.create(
                number=1, template=project.client.page_layout_templates.get(default=True))
            page.create_cells()

        parser_log.append(
            f"<div class='font-weight-bold'><hr>Totale prodotti importati: {project.products.count()}<hr></div>")

        project.excel_import_log = " ".join(parser_log)
        project.save()
    except Exception as error:
        project.excel_import_log = str(error)
        project.excel_import_failed = True
        project.save()
