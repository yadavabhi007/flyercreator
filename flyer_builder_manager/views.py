import base64
import logging
import subprocess
from io import BytesIO

import json
import requests
from django.db.models import Q
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core import files
from django.core.files.base import ContentFile
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q
from django.http import JsonResponse
from django.shortcuts import render, redirect,HttpResponse
from django.utils import timezone
from django.utils.formats import sanitize_separators
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from openpyxl import load_workbook

from .pagination import custom_pagination
from flyer_api.models import Project as Flyer_project, ProjectProductStyle, ProjectPageStyle, SpecialImage, ProjectPage, \
    ProjectStopper, ProjectPoster
from flyer_api.models import ProjectImage, ProjectProduct, ProjectPageCell, BannerImage, Client, ProjectPageSpecialCell, \
    PageFormat, Category, PageLayoutTemplate,TemplateTheme,PageTemplate

logger = logging.getLogger(__name__)


@login_required
def project_list(request):
    if not request.user.is_superuser and request.user.profile.client.only_catalog:
        return redirect('flyer_builder_manager:catalog_products')
    if request.user.is_superuser:
        projects = Flyer_project.objects.order_by('-created_at').all()
        working_projects = Flyer_project.objects.filter(status=0).order_by('-updated_at').all()
        completed_projects = Flyer_project.objects.exclude(status=0).order_by('-updated_at').all()
        projects_count_where_initialization_in_progress = Flyer_project.objects.filter(status=0,
                                                                                       initialization_in_progress=True).count()
        context = {"projects": projects, "working_projects": working_projects, "completed_projects": completed_projects,
                   "projects_count_where_initialization_in_progress": projects_count_where_initialization_in_progress}
        return render(request, 'flyer_builder_manager/projects.html', context)
    projects = request.user.profile.client.projects.order_by(
        '-created_at').all()
    working_projects = request.user.profile.client.projects.filter(
        status=0).order_by('-updated_at').all()
    completed_projects = request.user.profile.client.projects.exclude(
        status=0).order_by('-updated_at').all()
    projects_count_where_initialization_in_progress = request.user.profile.client.projects.filter(
        status=0, initialization_in_progress=True).count()
    flyer_format = PageFormat.objects.filter(type="flyer")
    stopper_format = PageFormat.objects.filter(type="stopper")
    poster_format = PageFormat.objects.filter(type="poster")
    number_of_products = PageLayoutTemplate.objects.filter(default=True)
    image_category = Category.objects.all()
    context = {"projects": projects, 'flyer_format': flyer_format, "stopper_format": stopper_format,
               "poster_format": poster_format, "working_projects": working_projects,'number_of_products':number_of_products,
               "completed_projects": completed_projects,
               "projects_count_where_initialization_in_progress": projects_count_where_initialization_in_progress,'image_category':image_category}
    return render(request, 'flyer_builder_manager/projects.html', context)

@login_required
def create_empty_project(request):
    data = eval(request.POST['data'])
    theme_id = request.POST['theme_id']
    theme=TemplateTheme.objects.get(id=theme_id)
    data_format=data["page_format"]
    page_format = theme.page_format
    stopper_format = PageFormat.objects.get(id=int(data["stopper_format"][0]))
    poster_format = PageFormat.objects.get(id=int(data["poster_format"][0]))
    number_of_products = data["number_of_products"][0]
    number_of_products = json.loads(number_of_products)
    rows = number_of_products['row']
    columns = number_of_products['columns']
    number_of_pages = int(data["number_of_pages"][0])
    category = theme.category
    page_templates=theme.templates.all()
    project = request.user.profile.client.projects.create(name=data["name"][0], page_format=page_format,
            stopper_format=stopper_format, poster_format=poster_format,category=category)
    page_style = ProjectPageStyle.objects.create()
    project.product_page_style = page_style
    project.save()
    no_of_product=PageLayoutTemplate.objects.create(name=f"{rows}x{columns}",rows=rows,columns=columns)
    for i in range(1, number_of_pages+1):
        if i ==1:
            page_template = page_templates.get(page_type='first_page')
        elif i == number_of_pages:
            page_template = page_templates.get(page_type='last_page')
        else:
            page_template = page_templates.get(page_type='inner_page')
        
        page_style = ProjectPageStyle.objects.create()
        page_style.header_per = page_template.header_per
        page_style.footer_per = page_template.footer_per
        page_style.body_per = (100 - (page_template.header_per + page_template.footer_per))
        
        if page_template.full_image: 
            page_style.header_per = 100
            page_style.body_per = 0
            page_style.footer_per = 0
        else:
            if page_template.half_image and page_template.page_type=="first_page":
                page_style.header_per = 50
            elif page_template.half_image and page_template.page_type=="last_page":
                page_style.footer_per = 50
        page_style.save() 
        page = project.pages.create(number=i, template=no_of_product, product_page_style = page_style)
        page.create_cells()
        page.create_super_cells()
        header_cell = page.special_cell.get(type='header')
        header_cell.image = page_template.header_image
        header_cell.save()
        footer_cell = page.special_cell.get(type='footer')
        footer_cell.image = page_template.footer_image
        footer_cell.save()
    return redirect('flyer_builder_manager:projects')



@login_required
def get_themes(request):
    page_format = PageFormat.objects.get(id=request.POST["page_format"])
    category = Category.objects.get(id=request.POST["category"])
    themes = TemplateTheme.objects.filter(category=category, page_format=page_format)
    context = {}
    context['data'] = dict(request.POST)
    context['themes'] = themes
    return render(request, 'flyer_builder_manager/_select_project_theme.html', context)

@login_required
@require_http_methods(["POST"])
@csrf_exempt
def create_excel_project(request):
    # project = request.user.projects.create(name=request.POST["name"])
    project = request.user.profile.client.projects.create(
        name=request.POST["name"])
    project.project_template_file = request.FILES['excel']
    project.initialization_in_progress = True
    project.save()
    subprocess.Popen([settings.PYTHON_PATH, settings.MANAGE_PY_PATH,
                      'initprojectbyexcel', str(project.id)], stdin=subprocess.PIPE)
    messages.success(
        request, 'Progetto inizializzato. Attendere per favore che la creazione sia completata')
    return redirect('flyer_builder_manager:projects')

    file_path = project.project_template_file.path
    workbook = load_workbook(filename=file_path, read_only=True)
    sheet = workbook.get_sheet_by_name(workbook.sheetnames[0])

    project_pages = {}
    current_page_number = 0
    # seller = Seller.objects.get(pk=request.user.profile.seller_code)
    seller = Seller.objects.get(pk=request.user.profile.client.seller_code)

    for index, row in enumerate(sheet.rows):
        if index >= 2:
            if row[0].value == "codice template":
                current_page_number = current_page_number + 1
                project_pages[current_page_number] = {
                    "has_products": False,
                    "template_code": row[3].value,
                    "products": []
                }
            elif isinstance(row[0].value, int) and (row[2].value is not None):
                project_pages[current_page_number]["has_products"] = True
                project_pages[current_page_number]["products"].append(
                    {
                        "code": row[2].value,
                        "position": row[1].value,
                        "other_codes": row[3].value,
                        "additional_data": {
                            "template_description": row[4].value
                        },
                        "note": row[5].value,
                        "price": row[6].value,
                        "price_without_discount": row[7].value,
                        "discount_percentage": ((row[8].value * 100) if (row[8].value is not None) else None),
                        "pieces_number": row[9].value,
                        "loyalty": (row[10].value == "x"),
                        "focus": (row[11].value == "x"),
                        "stopper": (row[12].value == "x"),
                        "poster": (row[13].value == "x"),
                    }
                )

    for page_number, excel_page_info in project_pages.items():
        if excel_page_info["has_products"]:
            # page = project.pages.create(
            #     number=page_number, template=PageLayoutTemplate.objects.filter(code=excel_page_info["template_code"])[0])
            page = project.pages.create(
                number=page_number, template=request.user.profile.client.page_layout_templates.filter(code=excel_page_info["template_code"])[0])
            page.create_cells()
            page_cells = page.cells.order_by(
                'y_top_left_coord', 'x_top_left_coord')
            current_page_free_cells = []
            for cell in page_cells:
                current_page_free_cells.append(cell)
            for excel_product in excel_page_info["products"]:
                distribution = Distribution.objects.get(
                    code=excel_product["code"], seller=seller)
                catalog_product = distribution.product
                current_cell = None
                if len(current_page_free_cells) > 0:
                    current_cell = current_page_free_cells.pop(0)
                product = project.products.create(
                    cell=current_cell,
                    ref_id=catalog_product.pk,
                    code=excel_product["code"],
                    description=catalog_product.get_description(),
                    description_brand=catalog_product.get_description_brand(),
                    description_type=catalog_product.get_description_type(),
                    description_tastes=catalog_product.get_description_tastes(),
                    description_weight=catalog_product.get_description_weight(),
                    position=excel_product["position"],
                    other_codes=excel_product["other_codes"],
                    additional_data=excel_product["additional_data"],
                    note=excel_product["note"],
                    price=excel_product["price"],
                    price_without_discount=excel_product["price_without_discount"],
                    discount_percentage=excel_product["discount_percentage"],
                    pieces_number=excel_product["pieces_number"],
                    loyalty=excel_product["loyalty"],
                    focus=excel_product["focus"],
                    stopper=excel_product["stopper"],
                    poster=excel_product["poster"]
                )
                catalog_picture = catalog_product.picture.filter(favorite=True)[
                    0]
                catalog_picture_url = settings.CATALOG_IMAGES_URL_PREFIX + \
                    catalog_picture.low_resolution_path()
                catalog_picture_image_response = requests.get(
                    catalog_picture_url)
                if catalog_picture_image_response.status_code == requests.codes.ok:
                    bytes_io = BytesIO()
                    bytes_io.write(catalog_picture_image_response.content)
                    file_name = catalog_picture_url.split("/")[-1]
                    product_image = product.images.create()
                    product_image.image.save(file_name, files.File(bytes_io))

    messages.success(request, 'Progetto creato')
    return redirect('flyer_builder_manager:projects')


@login_required
def new_project_page(request, project_id):
    project = request.user.profile.client.projects.get(pk=project_id)
    project_pages_count = project.pages.count()
    default_page_template = request.user.profile.client.page_layout_templates.filter(
        default=True)[0]

    start = int(request.GET["start"])
    position = request.GET["position"]

    if position == "next":
        pages_to_shift = project.pages.filter(
            number__gt=start).order_by('-number')
        for page_to_shift in pages_to_shift:
            page_to_shift.number += 1
            page_to_shift.save()
        new_page_number = start + 1
    if position == "prev":
        pages_to_shift = project.pages.filter(
            number__gte=start).order_by('-number')
        for page_to_shift in pages_to_shift:
            page_to_shift.number += 1
            page_to_shift.save()
        new_page_number = start

    page = project.pages.create(
        number=new_page_number, template=default_page_template)
    
    header_banner=ProjectPageSpecialCell(page = page,type='header')
    header_banner.save()
    footer_banner=ProjectPageSpecialCell(page = page,type='footer')
    footer_banner.save()
    page.create_cells()
    project.update_last_modification_time()
    messages.success(request, 'Pagina creata')
    return redirect('flyer_builder_manager:edit_project_page', project_id=project.pk, page_number=page.number)


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def delete_project_product(request, project_id, product_id):
    # project = request.user.projects.get(pk=project_id)
    project = request.user.profile.client.projects.get(pk=project_id)
    product = ProjectProduct.objects.filter(project=project).get(pk=product_id)
    product.delete()
    project.update_last_modification_time()
    return JsonResponse({"status": "deleted"})


@login_required
# @require_http_methods(["POST"])
# @csrf_exempt
def delete_project_page(request, project_id, page_number):
    # project = request.user.projects.get(pk=project_id)
    project = request.user.profile.client.projects.get(pk=project_id)
    delete_products = True if (
        request.GET["delete_products"] == "yes") else False
    if project.pages.count() > 1:
        is_last_page = (project.pages.count() == page_number)
        page = project.pages.filter(number=page_number)[0]
        page_cells = page.cells.all()
        if delete_products:
            for page_cell in page_cells:
                if ProjectProduct.objects.filter(cell=page_cell).count() > 0:
                    product = ProjectProduct.objects.filter(cell=page_cell)[0]
                    product.delete()
        page.delete()
        pages_to_change = project.pages.filter(number__gte=page_number)
        for page_to_change in pages_to_change:
            page_to_change.number = page_to_change.number - 1
            page_to_change.save()
        project.update_last_modification_time()
        messages.success(request, 'Pagina cancellata')
        return redirect('flyer_builder_manager:edit_project_page', project_id=project.pk, page_number=((page_number - 1) if is_last_page else page_number))
    messages.error(
        request, '<h3>ERRORE</h3> <p>Non è possibile cancellare la pagina: il volantino ne deve contenere almeno una!</p>')
    return redirect('flyer_builder_manager:edit_project_page', project_id=project_id, page_number=page_number)


@login_required
def clear_project_page(request, project_id, page_number):
    # project = request.user.projects.get(pk=project_id)
    project = request.user.profile.client.projects.get(pk=project_id)
    delete_products = True if (
        request.GET["delete_products"] == "yes") else False
    page = project.pages.filter(number=page_number)[0]
    page_cells = page.cells.all()
    for page_cell in page_cells:
        if page_cell.has_product():
            product = page_cell.product()
            if delete_products:
                product.delete()
            else:
                product.cell = None
                product.save()
    project.update_last_modification_time()
    messages.success(request, 'Pagina pulita')
    return redirect('flyer_builder_manager:edit_project_page', project_id=project_id, page_number=page_number)


@login_required
def switch_project_page(request, project_id, page_number):
    # project = request.user.projects.get(pk=project_id)
    project = request.user.profile.client.projects.get(pk=project_id)
    page_number_to_switch = request.GET["with"]
    page_1 = project.pages.filter(number=page_number)[0]
    page_2 = project.pages.filter(number=page_number_to_switch)[0]
    page_number_1 = page_1.number
    page_number_2 = page_2.number
    page_1.number = 0
    page_1.save()
    page_2.number = page_number_1
    page_2.save()
    page_1.number = page_number_2
    page_1.save()
    project.update_last_modification_time()
    messages.success(request, 'Pagine scambiate')
    return redirect('flyer_builder_manager:edit_project_page', project_id=project_id, page_number=page_number)


@login_required
def project_page_change_template(request, project_id, page_number, template_id):
    # project = request.user.projects.get(pk=project_id)
    project = request.user.profile.client.projects.get(pk=project_id)
    page = project.pages.filter(number=page_number)[0]
    # template = PageLayoutTemplate.objects.get(pk=template_id)
    template = request.user.profile.client.page_layout_templates.get(
        pk=template_id)
    cells = page.cells.all()
    products_to_migrate = []
    for cell in cells:
        if cell.has_product():
            products_to_migrate.append(cell.product())
        cell.delete()
    page.template = template
    page.save()
    page.create_cells()
    new_cells = page.cells.order_by('y_top_left_coord', 'x_top_left_coord')
    for idx, cell in enumerate(new_cells):
        if idx < len(products_to_migrate):
            products_to_migrate[idx].cell = cell
            products_to_migrate[idx].save()
    project.update_last_modification_time()
    return redirect('flyer_builder_manager:edit_project_page', project_id=project_id, page_number=page_number)


@login_required
def delete_project(request, uid):
    # project = request.user.projects.get(pk=uid)
    project = request.user.profile.client.projects.get(pk=uid)
    project.products.all().delete()
    project.delete()
    messages.success(request, 'Progetto eliminato')
    return redirect('flyer_builder_manager:projects')


@login_required
def search_product(request):
    code = request.GET["code"]
    description = request.GET["description"]
    seller_id = request.user.profile.client.seller_code
    # prods = Product.objects
    # if code != '':
    #     prods = prods.filter(product_id__in=Distribution.objects.filter(
    #         code__icontains=code).values('product_id'))
    # prods = prods.filter(product_id__in=Distribution.objects.filter(
    #     seller__id=seller_id).values('product_id'))
    # if description != '':
    #     prods = prods.filter(product_id__in=Description.objects.filter(
    #         Q(description__icontains=description)).values('product_id'))
    # products = prods.order_by('-lastModify')[:10]
    context = {"products": [{
        'pk':1,
        'image_url':'/media/banner/timer-banner.jpg',
        'get_code':'POPOPO',
        'get_description':'Descrizione'
    },
    {
        'pk':2,
        'image_url':'/media/banner/2016-04-20-banner-popup-advertising.jpg',
        'get_code':'pro2',
        'get_description':'Descrizione 2'
    },
    {
        'pk':3,
        'image_url':'/media/banner/coca-cola-transparent-png-pictures-icons-and-png-25.png',
        'get_code':'cccl',
        'get_description':'Coca Cola'
    }
    ]}
    return render(request, 'flyer_builder_manager/_search_product_result.html', context)


@login_required
def subcategory_select_options(request):
    category_id = request.GET["category_id"]
    subcategories = Subcategory.objects.filter(category=category_id)
    context = {'subcategories': subcategories}
    return render(request, 'flyer_builder_manager/product/_subcategory_select_options.html', context)


@login_required
def edit_project_page(request, project_id, page_number):
    # project = request.user.projects.get(pk=project_id)
    # import pdb;pdb.set_trace()
    project = request.user.profile.client.projects.get(pk=project_id)
    page = project.pages.filter(number=page_number)[0]
    category = project.category
    default_special_image = SpecialImage.objects.filter(category=category, client=None)
    header_banner = ProjectPageSpecialCell.objects.get(page=page,type='header')
    footer_banner = ProjectPageSpecialCell.objects.get(page=page,type='footer')
    user_header_images = request.user.profile.client.specialimage_set.filter(type='header', category=category)
    header_default = default_special_image.filter(type='header')
    # header_images = header_default|user_header_images
    user_footer_images = request.user.profile.client.specialimage_set.filter(type='footer', category=category)
    footer_default = default_special_image.filter(type='footer')
    # footer_images = footer_default|user_footer_images
    user_full_page_images = request.user.profile.client.specialimage_set.filter(type='full_page',category=category)
    full_page_default = default_special_image.filter(type='full_page')
    # full_page_images = full_page_default|user_full_page_images
    user_half_page_images = request.user.profile.client.specialimage_set.filter(type='half_page',category=category)
    half_page_default = default_special_image.filter(type='half_page')
    # half_page_images = half_page_default|user_half_page_images
    available_products = ProjectProduct.objects.filter(
        project=project, cell=None)
    render_sizes = {
        "page_width": int(page.get_page_render_width() * request.user.profile.zoom_factor()),
        "page_height": int(page.get_page_render_height() * request.user.profile.zoom_factor())
    }
    image_category = Category.objects.all()
    context = {'project': project, 'page': page,
               'available_products': available_products,
               'render_sizes': render_sizes,
               'header_banner':header_banner,
               'footer_banner':footer_banner,
               "header_images":user_header_images,
               "header_default":header_default,
               "footer_images":user_footer_images,
               "footer_default":footer_default,
               "full_page_images":user_full_page_images,
               "full_page_default":full_page_default,
               "half_page_images":user_half_page_images,
               "half_page_default":half_page_default,
               "image_category":image_category
               }
    return render(request, 'flyer_builder_manager/project/page/edit.html', context)


@login_required
def project_page_partial(request, project_id, page_number):
    # project = request.user.projects.get(pk=project_id)
    project = request.user.profile.client.projects.get(pk=project_id)
    page = project.pages.filter(number=page_number)[0]
    render_sizes = {
        "page_width": int(page.get_page_render_width() * request.user.profile.zoom_factor()),
        "page_height": int(page.get_page_render_height() * request.user.profile.zoom_factor())
    }
    # header_banner=ProjectPageSpecialCell.objects.get(page=page,type='header')
    # footer_banner=ProjectPageSpecialCell.objects.get(page=page,type='footer')
    # header_images = request.user.profile.client.specialimage_set.filter(type='header')
    # footer_images = request.user.profile.client.specialimage_set.filter(type='footer')
    # full_page_images = request.user.profile.client.specialimage_set.filter(type='full_page')
    # half_page_images = request.user.profile.client.specialimage_set.filter(type='half_page')
    
    
    category = project.category
    default_special_image = SpecialImage.objects.filter(category=category, client=None)
    header_banner = ProjectPageSpecialCell.objects.get(page=page,type='header')
    footer_banner = ProjectPageSpecialCell.objects.get(page=page,type='footer')
    user_header_images = request.user.profile.client.specialimage_set.filter(type='header',category=category)
    header_default = default_special_image.filter(type='header')
    header_images = header_default|user_header_images
    user_footer_images = request.user.profile.client.specialimage_set.filter(type='footer', category=category)
    footer_default = default_special_image.filter(type='footer')
    footer_images = footer_default|user_footer_images
    user_full_page_images = request.user.profile.client.specialimage_set.filter(type='full_page',category=category)
    full_page_default = default_special_image.filter(type='full_page')
    full_page_images = full_page_default|user_full_page_images
    user_half_page_images = request.user.profile.client.specialimage_set.filter(type='half_page',category=category)
    half_page_default = default_special_image.filter(type='half_page')
    half_page_images = half_page_default|user_half_page_images
    
    context = {'project': project, 'page': page,
               'render_sizes': render_sizes,
               'header_banner':header_banner,
               'footer_banner':footer_banner,
               "header_images":header_images,
               "footer_images":footer_images,
               "full_page_images":full_page_images,
               "half_page_images":half_page_images,
               }
    return render(request, 'flyer_builder_manager/project/page/_page.html', context)


@login_required
def project_available_products_partial(request, project_id):
    # project = request.user.projects.get(pk=project_id)
    project = request.user.profile.client.projects.get(pk=project_id)
    available_products = ProjectProduct.objects.filter(
        project=project, cell=None)
    context = {'project': project, 'available_products': available_products}
    return render(request, 'flyer_builder_manager/project/page/_available_products.html', context)


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def import_catalog_product_to_cell(request):
    products=[{
        'pk':1,
        'price':1.20,
        'image_url':'/media/banner/timer-banner.jpg',
        'get_code':'POPOPO',
        'get_description':'Descrizione'
    },
    {
        'pk':2,
        'price':2.30,
        'image_url':'https://www.artera.net/wp-content/uploads/2016/04/2016-04-20-banner-popup-advertising.jpg',
        'get_code':'pro2',
        'get_description':'Descrizione 2'
    },
    {
        'pk':3,
        'price':2.00,
        'image_url':'/media/banner/coca-cola-transparent-png-pictures-icons-and-png-25.png',
        'get_code':'cccl',
        'get_description':'Coca Cola'
    }
    ]
    
    cell_id = request.POST["cell_id"]
    catalog_product_id = request.POST["catalog_product_id"]
    
    for pro in products:
        if int(pro['pk']) == int(catalog_product_id):
            choised_product = pro
    # catalog_product = Product.objects.get(pk=catalog_product_id)
    cell = ProjectPageCell.objects.get(pk=cell_id)
    project = cell.page.project
    # catalog_product_code = Distribution.objects.filter(
    #     product=catalog_product)[0].code
    
    product = project.products.create(
        price=choised_product['price'],
        cell=cell,
        # id prodotto
        ref_id=choised_product['pk'],
        #codice prodotto
        code=choised_product['get_code'],
        description=choised_product["get_description"],
        description_brand='Descrizione 1',
        description_type='Descrizione 2',
        description_tastes='Descrizione 3',
        description_weight='Descrizione 4',
        
    )
    image= ProjectImage(project_product=product, url=choised_product['image_url'])
    image.save()
    
    # catalog_picture = catalog_product.picture.filter(favorite=True)[0]
    # catalog_picture = 'ciao.jpg'
    # catalog_picture_url = settings.CATALOG_IMAGES_URL_PREFIX + \
    #     catalog_picture.low_resolution_path()
    # catalog_picture_url = 'http://pbs.twimg.com/profile_images/758084549821730820/_HYHtD8F.jpg'
    # catalog_picture_image_response = requests.get(catalog_picture_url)
    # if catalog_picture_image_response.status_code == requests.codes.ok:
    #     bytes_io = BytesIO()
    #     bytes_io.write(catalog_picture_image_response.content)
    #     file_name = catalog_picture_url.split("/")[-1]
    #     product_image = product.images.create()
    #     product_image.image.save(file_name, files.File(bytes_io))
    project.update_last_modification_time()
    return JsonResponse({"status": "imported"})


@login_required
def get_project_product_data(request, project_id, product_id):
    project = request.user.profile.client.projects.get(pk=project_id)
    product = project.products.get(pk=product_id)
    block_image_url = False
    if product.ref_block_id is not None:
        try:
            block = Block.objects.get(pk=product.ref_block_id)
            block_image_url = block.preview_path()
        except:
            block_image_url = False
    return JsonResponse({
        "price": product.price,
        "description1": product.description_brand,
        "description2": product.description_type,
        "description3": product.description_tastes,
        "description4": product.description_weight,
        "loyalty": product.loyalty,
        "focus": product.focus,
        "stopper": product.stopper,
        "poster": product.poster,
        "ref_block_id": product.ref_block_id,
        "block_image_url": block_image_url,
        "note": product.note,
        "price_without_discount": product.price_without_discount,
        "tag_code": product.tag_code,
        "discount_percentage": product.discount_percentage,
        "pieces_number": product.pieces_number,
        "max_purchasable_pieces": product.max_purchasable_pieces,
        "points": product.points,
    })


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def edit_project_product_data(request, project_id, product_id):
    project = request.user.profile.client.projects.get(pk=project_id)
    product = project.products.get(pk=product_id)
    try:
        price = float(sanitize_separators(request.POST["price"]))
    except:
        price = float(0)
    try:
        price_without_discount = float(sanitize_separators(
            request.POST["price_without_discount"]))
    except:
        price_without_discount = float(0)
    try:
        discount_percentage = float(sanitize_separators(
            request.POST["discount_percentage"]))
    except:
        discount_percentage = float(0)
    try:
        pieces_number = int(sanitize_separators(request.POST["pieces_number"]))
    except:
        pieces_number = int(0)
    try:
        max_purchasable_pieces = int(sanitize_separators(
            request.POST["max_purchasable_pieces"]))
    except:
        max_purchasable_pieces = int(0)
    try:
        points = int(sanitize_separators(request.POST["points"]))
    except:
        points = int(0)
    description1 = request.POST["description1"]
    description2 = request.POST["description2"]
    description3 = request.POST["description3"]
    description4 = request.POST["description4"]
    tag_code = request.POST["tag_code"]
    note = request.POST["note"]
    loyalty = request.POST.get("loyalty", None)
    if loyalty is not None:
        loyalty = True if (request.POST["loyalty"] == "true") else False
        product.loyalty = loyalty
    focus = request.POST.get("focus", None)
    if focus is not None:
        focus = True if (request.POST["focus"] == "true") else False
        product.focus = focus
    stopper = True if (request.POST["stopper"] == "true") else False
    poster = True if (request.POST["poster"] == "true") else False
    # if request.POST["ref_block_id"]:
    #     ref_block_id = request.POST["ref_block_id"]
    # else:
    ref_block_id = None
    product.price = price
    product.description_brand = description1
    product.description_type = description2
    product.description_tastes = description3
    product.description_weight = description4
    product.stopper = stopper
    product.poster = poster
    product.ref_block_id = ref_block_id
    product.note = note
    product.price_without_discount = price_without_discount
    product.tag_code = tag_code
    product.discount_percentage = discount_percentage
    product.pieces_number = pieces_number
    product.max_purchasable_pieces = max_purchasable_pieces
    product.points = points
    product.save()
    product.project.update_last_modification_time()
    return JsonResponse({
        "status": "edited",
    })


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def swap_cells(request):
    source_cell_id = request.POST["source_cell_id"]
    destination_cell_id = request.POST["destination_cell_id"]
    source_product_id = request.POST["source_product_id"]

    if int(source_cell_id) == 0:
        source_cell = None
    else:
        source_cell = ProjectPageCell.objects.get(pk=source_cell_id)

    if int(destination_cell_id) == 0:
        destination_cell = None
    else:
        destination_cell = ProjectPageCell.objects.get(pk=destination_cell_id)
        if ProjectProduct.objects.filter(cell=destination_cell_id).count() > 0:
            destination_cell_product = ProjectProduct.objects.filter(
                cell=destination_cell_id)[0]
        else:
            destination_cell_product = None

    if source_cell is not None:
        source_cell_product = ProjectProduct.objects.filter(cell=source_cell_id)[
            0]
    else:
        source_cell_product = None

    if (source_cell is not None) and (destination_cell is not None):
        source_cell_product.cell = destination_cell
        source_cell_product.save()
        if destination_cell_product is not None:
            destination_cell_product.cell = source_cell
            destination_cell_product.save()

    if (source_cell is None) and (destination_cell is not None):
        source_product = ProjectProduct.objects.get(pk=source_product_id)
        source_product.cell = destination_cell
        source_product.save()
        if destination_cell_product is not None:
            destination_cell_product.cell = None
            destination_cell_product.save()

    if (source_cell is not None) and (destination_cell is None):
        source_cell_product.cell = None
        source_cell_product.save()

    # project.update_last_modification_time()
    return JsonResponse({
        "status": "swapped",
    })


@login_required
def merge_split_cells(request, project_id, page_number, cell_id, operation):
    # project = request.user.projects.get(pk=project_id)
    project = request.user.profile.client.projects.get(pk=project_id)
    page = project.pages.filter(number=page_number)[0]
    target_cell = page.cells.get(pk=cell_id)

    if operation == "merge_left":
        if page.cells.filter(x_top_left_coord=(
                target_cell.x_top_left_coord - 1), y_top_left_coord=target_cell.y_top_left_coord).exists():
            cell_to_merge = page.cells.filter(x_top_left_coord=(
                target_cell.x_top_left_coord - 1), y_top_left_coord=target_cell.y_top_left_coord)[0]
        else:
            cell_to_merge = page.cells.filter(x_top_left_coord=(
                target_cell.x_top_left_coord - 2), y_top_left_coord=target_cell.y_top_left_coord)[0]

        if target_cell.has_product():
            new_cell_product = target_cell.product()
        elif cell_to_merge.has_product():
            new_cell_product = cell_to_merge.product()
        else:
            new_cell_product = None
        new_cell = page.cells.create()
        new_cell.x_top_left_coord = cell_to_merge.x_top_left_coord
        new_cell.y_top_left_coord = cell_to_merge.y_top_left_coord
        new_cell.width = target_cell.width + cell_to_merge.width
        new_cell.height = target_cell.height
        new_cell.save()
        if new_cell_product is not None:
            new_cell_product.cell = new_cell
            new_cell_product.save()
        target_cell.delete()
        cell_to_merge.delete()

    if operation == "merge_right":
        cell_to_merge = page.cells.filter(x_top_left_coord=(
            target_cell.x_top_left_coord + 1), y_top_left_coord=target_cell.y_top_left_coord)[0]
        if target_cell.has_product():
            new_cell_product = target_cell.product()
        elif cell_to_merge.has_product():
            new_cell_product = cell_to_merge.product()
        else:
            new_cell_product = None
        new_cell = page.cells.create()
        new_cell.x_top_left_coord = target_cell.x_top_left_coord
        new_cell.y_top_left_coord = target_cell.y_top_left_coord
        new_cell.width = target_cell.width + cell_to_merge.width
        new_cell.height = target_cell.height
        new_cell.save()
        if new_cell_product is not None:
            new_cell_product.cell = new_cell
            new_cell_product.save()
        target_cell.delete()
        cell_to_merge.delete()

    if operation == "merge_bottom":
        cell_to_merge = page.cells.filter(x_top_left_coord=target_cell.x_top_left_coord,
                                          y_top_left_coord=(target_cell.y_top_left_coord + 1))[0]
        if target_cell.has_product():
            new_cell_product = target_cell.product()
        elif cell_to_merge.has_product():
            new_cell_product = cell_to_merge.product()
        else:
            new_cell_product = None
        new_cell = page.cells.create()
        new_cell.x_top_left_coord = target_cell.x_top_left_coord
        new_cell.y_top_left_coord = target_cell.y_top_left_coord
        new_cell.width = target_cell.width
        new_cell.height = target_cell.height + cell_to_merge.height
        new_cell.save()
        if new_cell_product is not None:
            new_cell_product.cell = new_cell
            new_cell_product.save()
        target_cell.delete()
        cell_to_merge.delete()

    if operation == "merge_top":
        cell_to_merge = page.cells.filter(x_top_left_coord=target_cell.x_top_left_coord,
                                          y_top_left_coord=(target_cell.y_top_left_coord - 1))[0]
        if target_cell.has_product():
            new_cell_product = target_cell.product()
        elif cell_to_merge.has_product():
            new_cell_product = cell_to_merge.product()
        else:
            new_cell_product = None
        new_cell = page.cells.create()
        new_cell.x_top_left_coord = target_cell.x_top_left_coord
        new_cell.y_top_left_coord = target_cell.y_top_left_coord - 1
        new_cell.width = target_cell.width
        new_cell.height = target_cell.height + cell_to_merge.height
        new_cell.save()
        if new_cell_product is not None:
            new_cell_product.cell = new_cell
            new_cell_product.save()
        target_cell.delete()
        cell_to_merge.delete()

    if operation == "split_vert_2":
        if target_cell.has_product():
            new_cell_product = target_cell.product()
        else:
            new_cell_product = None
        width_new_cells = target_cell.width / 2
        new_cell_1 = page.cells.create()
        new_cell_1.x_top_left_coord = target_cell.x_top_left_coord
        new_cell_1.y_top_left_coord = target_cell.y_top_left_coord
        new_cell_1.width = width_new_cells
        new_cell_1.height = target_cell.height
        new_cell_1.save()
        new_cell_2 = page.cells.create()
        new_cell_2.x_top_left_coord = (
            target_cell.x_top_left_coord + 1)
        new_cell_2.y_top_left_coord = target_cell.y_top_left_coord
        new_cell_2.width = width_new_cells
        new_cell_2.height = target_cell.height
        new_cell_2.save()
        if new_cell_product is not None:
            new_cell_product.cell = new_cell_1
            new_cell_product.save()
        target_cell.delete()

    if operation == "split_vert_3":
        if target_cell.has_product():
            new_cell_product = target_cell.product()
        else:
            new_cell_product = None
        width_new_cells = target_cell.width / 3
        new_cell_1 = page.cells.create()
        new_cell_1.x_top_left_coord = target_cell.x_top_left_coord
        new_cell_1.y_top_left_coord = target_cell.y_top_left_coord
        new_cell_1.width = width_new_cells
        new_cell_1.height = target_cell.height
        new_cell_1.save()
        new_cell_2 = page.cells.create()
        new_cell_2.x_top_left_coord = (
            target_cell.x_top_left_coord + 1)
        new_cell_2.y_top_left_coord = target_cell.y_top_left_coord
        new_cell_2.width = width_new_cells
        new_cell_2.height = target_cell.height
        new_cell_2.save()
        new_cell_3 = page.cells.create()
        new_cell_3.x_top_left_coord = (
            target_cell.x_top_left_coord + (1 * 2))
        new_cell_3.y_top_left_coord = target_cell.y_top_left_coord
        new_cell_3.width = width_new_cells
        new_cell_3.height = target_cell.height
        new_cell_3.save()
        if new_cell_product is not None:
            new_cell_product.cell = new_cell_1
            new_cell_product.save()
        target_cell.delete()

    if operation == "split_horiz_2":
        if target_cell.has_product():
            new_cell_product = target_cell.product()
        else:
            new_cell_product = None
        height_new_cells = target_cell.height / 2
        new_cell_1 = page.cells.create()
        new_cell_1.x_top_left_coord = target_cell.x_top_left_coord
        new_cell_1.y_top_left_coord = target_cell.y_top_left_coord
        new_cell_1.width = target_cell.width
        new_cell_1.height = height_new_cells
        new_cell_1.save()
        new_cell_2 = page.cells.create()
        new_cell_2.x_top_left_coord = target_cell.x_top_left_coord
        new_cell_2.y_top_left_coord = (
            target_cell.y_top_left_coord + 1)
        new_cell_2.width = target_cell.width
        new_cell_2.height = height_new_cells
        new_cell_2.save()
        if new_cell_product is not None:
            new_cell_product.cell = new_cell_1
            new_cell_product.save()
        target_cell.delete()

    page.set_cells_position()
    project.update_last_modification_time()
    return redirect('flyer_builder_manager:edit_project_page', project_id=project.pk, page_number=page.number)


@login_required
def zoom(request, project_id, page_number):
    # project = request.user.projects.get(pk=project_id)
    project = request.user.profile.client.projects.get(pk=project_id)
    zoom_type = request.GET["type"]
    if zoom_type == "in":
        if request.user.profile.project_page_zoom < 110:
            request.user.profile.project_page_zoom = request.user.profile.project_page_zoom + 10
            request.user.profile.save()
    if zoom_type == "out":
        if request.user.profile.project_page_zoom > 70:
            request.user.profile.project_page_zoom = request.user.profile.project_page_zoom - 10
            request.user.profile.save()
    return redirect('flyer_builder_manager:edit_project_page', project_id=project.pk, page_number=page_number)


@login_required
@require_http_methods(["POST", "GET"])
@csrf_exempt
def generate_pdf(request, project_id):
    
    project = request.user.profile.client.projects.get(pk=project_id)
    if request.method=="GET":
        pages = project.pages.all().order_by("number")
        images = []
        for page in pages:
            page_image_path = page.image.url
            images.append(page_image_path)
        context = {'project': project, 'images': images, 'width': int(project.page_format.width),
                   "height": int(project.page_format.height)}
        return render(request, 'flyer_builder_manager/project/pdf.html', context)
    project.pdf_generation_in_progress = True
    project.save()
    pdf_type= request.POST.get("type", 'page')
    subprocess.Popen([settings.PYTHON_PATH, settings.MANAGE_PY_PATH, 'generateprojectpdf', str(
        project.id), str(request.build_absolute_uri()), pdf_type], stdin=subprocess.PIPE)
    return JsonResponse({
        "status": "generating",
    })

import pdfkit
from flyer_api.models import Project
from django.template.loader import render_to_string

@login_required
def test_pdf(request, project_id):
    if request.is_secure():
        site_url = 'https://' + request.get_host()
    else:
        site_url = 'http://' + request.get_host()
    pdf_type= request.POST.get("type", 'page')
    project = request.user.profile.client.projects.get(pk=project_id)
    return generatepdf( project.id,str(request.build_absolute_uri()),pdf_type,site_url)

    
def generatepdf(project_id,absolute_uri,pdf_type,site_url):
    print("HIIIIII")
    project = Project.objects.get(pk=project_id)
    images = []
    pdf_page_width = 0
    pdf_page_height = 0
    if pdf_type == "page":
        pages = project.pages.all().order_by("number")
        for page in pages:
            page_image_path = page.image.url
            images.append(page_image_path)
            pdf_page_width = int(project.page_format.width)
            pdf_page_height = int(project.page_format.height)
    elif pdf_type == "stopper":
        stoppers = project.stoppers.all()
        for stopper in stoppers:
            image_path = stopper.image.url
            images.append(image_path)
            pdf_page_width = int(project.stopper_format.width)
            pdf_page_height = int(project.stopper_format.height)
    elif pdf_type == "poster":
        posters = project.posters.all()
        for poster in posters:
            image_path = poster.image.url
            images.append(image_path)
            pdf_page_width = int(project.poster_format.width)
            pdf_page_height = int(project.poster_format.height)
    print(len(images))
    context = {'project': project, 'images': images, 'width':pdf_page_width, "height": pdf_page_height,'site_url':site_url}
    html = render_to_string("flyer_builder_manager/project/pdf.html", context=context)
    options = {
        'page-size':'A4',
        'encoding':'utf-8', 
        'margin-top':'0.1in',
        'margin-bottom':'0.1in',
        'margin-left':'0.1in',
        'margin-right':'0.1in'
    }
    print("pdf_page_width",pdf_page_width)
    print("pdf_page_height",pdf_page_height)

    output = pdfkit.from_string(html, output_path=False,options=options)
    response = HttpResponse(content_type="application/pdf")
    response.write(output)
    # bytes_io = BytesIO(output)
    # if pdf_type == "stopper":
    #     project.project_stopper_pdf_file.save(f"{project.name}_stopper.pdf", content=bytes_io)
    # elif pdf_type == "poster":
    #     project.project_poster_pdf_file.save(f"{project.name}_poster.pdf", content=bytes_io)
    # else:
    #     project.project_pdf_file.save(f"{project.name}.pdf", content=bytes_io)
    project.pdf_last_generation = timezone.now()
    project.pdf_generation_in_progress = False
    project.save()
    return response
    

@login_required
def pdf_generating_status(request, project_id, pdf_type):
    project = request.user.profile.client.projects.get(pk=project_id)
    if project.pdf_generation_in_progress:
        status = "generation_in_progress"
        pdf_generation_time = None
        pdf_url = None
    else:
        status = "generated"
        pdf_generation_time = timezone.localtime(
            project.pdf_last_generation).strftime("%d/%m/%Y %H:%M")
        if pdf_type =="page":
            pdf_url = project.project_pdf_file.url
        elif pdf_type =="stopper":
            pdf_url = project.project_stopper_pdf_file.url
        elif pdf_type =="poster":
            pdf_url = project.project_poster_pdf_file.url
    return JsonResponse({
        "status": status,
        "pdf_generation_time": pdf_generation_time,
        "pdf_url": pdf_url
    })


@login_required
def initialization_status(request, project_id):
    project = request.user.profile.client.projects.get(pk=project_id)
    if project.initialization_in_progress:
        status = "initialization_in_progress"
    else:
        if project.excel_import_failed:
            status = "excel_import_failed"
        else:
            status = "initialization_completed"
    return JsonResponse({
        "status": status
    })


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def update_products_data_by_excel(request, project_id):
    request.upload_handlers.pop(0)
    project = request.user.profile.client.projects.get(pk=project_id)
    excel_file = request.FILES['excel']
    file_path = excel_file.temporary_file_path()
    workbook = load_workbook(filename=file_path, read_only=True)
    sheet = workbook.get_sheet_by_name(workbook.sheetnames[0])
    updated_product_count = 0
    not_found_products = 0

    for index, row in enumerate(sheet.rows):
        if index > 0:
            # if (row[0].value is not None) and (row[1].value is not None):
            if (row[0].value is not None):
                product = None
                try:
                    product = project.products.filter(code=row[0].value)[0]
                except:
                    not_found_products += 1
                if product is not None:
                    if (row[1].value is not None):
                        product.price = row[1].value
                    if (row[2].value is not None):
                        product.price_without_discount = row[2].value
                    if (row[3].value is not None):
                        product.discount_percentage = (row[3].value * 100)
                    if (row[4].value is not None):
                        product.pieces_number = row[4].value
                    if (row[5].value is not None):
                        product.max_purchasable_pieces = row[5].value
                    if (row[6].value is not None):
                        product.points = row[6].value
                    if (row[7].value is not None):
                        if row[7].value.lower() == "si":
                            product.stopper = True
                        elif row[7].value.lower() == "no":
                            product.stopper = False
                    if (row[8].value is not None):
                        if row[8].value.lower() == "si":
                            product.poster = True
                        elif row[8].value.lower() == "no":
                            product.poster = False
                    product.save()
                    updated_product_count += 1

    feedback_message = f"<strong>Aggiornamento dati completato</strong>:<br> {updated_product_count} prodotti aggiornati"
    if not_found_products > 0:
        feedback_message += f"<br> {not_found_products} prodotti non trovati"

    messages.success(request, feedback_message)

    return redirect('flyer_builder_manager:edit_project_page', project_id=project.pk, page_number=1)


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def change_page_name(request, project_id, page_number):
    name = request.POST['name']
    project = request.user.profile.client.projects.get(pk=project_id)
    page = project.pages.filter(number=page_number)[0]
    page.name = name
    page.save()
    return JsonResponse({
        "status": "name updated"
    })


@login_required
def send_to_agency(request, project_id):
    project = request.user.profile.client.projects.get(pk=project_id)

    seller = Seller.objects.get(pk=request.user.profile.client.seller_code)
    layout = Layout.objects.get(pk=2)
    aeb_project = Project.objects.create(
        name=project.name,
        sell_in=timezone.now(),
        sell_out=timezone.now(),
        seller=seller,
        # layout=[layout],
        project_type=1,
        default_flyer_block=Block.objects.get(
            pk=request.user.profile.client.flyer_block_id_ref),
        default_locandina_block=Block.objects.get(
            pk=request.user.profile.client.locandina_block_id_ref),
        default_stopper_block=Block.objects.get(
            pk=request.user.profile.client.stopper_block_id_ref)
    )
    aeb_project.layout.add(layout)

    for page in project.pages.order_by("number"):
        page_name = "Pagina" if page.name is None else page.name
        aeb_page = ProjPage.objects.create(
            project=aeb_project,
            pageNum=page.number,
            name=page_name
        )

    for page in project.pages.order_by("number"):
        page_cells = page.cells.all()
        for page_cell in page_cells:
            if page_cell.has_product():
                product = page_cell.product()
                aeb_product = Product.objects.get(pk=product.ref_id)
                default_block = Block.objects.get(
                    pk=request.user.profile.client.flyer_block_id_ref)
                default_locandina_block = Block.objects.get(
                    pk=request.user.profile.client.locandina_block_id_ref)
                default_stopper_block = Block.objects.get(
                    pk=request.user.profile.client.stopper_block_id_ref)
                product_points = "" if product.points is None else product.points
                product_note = "" if product.note is None else product.note
                # product_price_without_discount = "" if product.price_without_discount is None else int(
                #     product.price_without_discount)
                product_price_without_discount = "" if product.price_without_discount is None else str(
                    product.price_without_discount)
                note3 = "-".join(filter(None,
                                        [product.tag_code, product_price_without_discount]))
                product_discount_percentage = "" if product.discount_percentage is None else int(
                    product.discount_percentage)
                aeb_item = Item.objects.create(
                    project=aeb_project,
                    product=aeb_product,
                    pageNum=page.number,
                    positionNum=product.cell.position,
                    flyer_block=default_block,
                    locandina_block=default_locandina_block,
                    stopper_block=default_stopper_block,
                    field1=product.description_brand,
                    field2=product.description_type,
                    field3=product.description_tastes,
                    field4=product.description_weight,
                    grammageValue=aeb_product.grammageValue,
                    category=aeb_product.category.name,
                    price_with_IVA=product.price,
                    onStopper=product.stopper,
                    onLocandina=product.poster,
                    note2=product_discount_percentage,
                    punti=product_points,
                    note=product_note,
                    available_pieces=product.pieces_number,
                    note3=note3,
                    max_purchasable_pieces=product.max_purchasable_pieces
                )
                if product.ref_block_id is not None:
                    block = Block.objects.get(pk=product.ref_block_id)
                    aeb_item.flyer_block = block
                    if aeb_item.onStopper:
                        if Block.objects.filter(seller=seller, code=f"{block.code}Stop").exists():
                            stopper_block = Block.objects.get(
                                seller=seller, code=f"{block.code}Stop")
                            aeb_item.stopper_block = stopper_block
                        if Block.objects.filter(seller=seller, code=f"{block.code}Loc").exists():
                            locandina_block = Block.objects.get(
                                seller=seller, code=f"{block.code}Loc")
                            aeb_item.locandina_block = locandina_block
                    aeb_item.save()


                aeb_pictures = aeb_product.picture.all()
                for aeb_picture in aeb_pictures:
                    pic_item = PictureItem(
                        picturePath=aeb_picture.path, orderNum=0)
                    pic_item.save()
                    aeb_item.picture_item.add(pic_item)
                    aeb_item.save()

    project.status = 1
    project.save()
    subprocess.Popen([settings.PYTHON_PATH, settings.MANAGE_PY_PATH, 'generateprojectpdf', str(
        project.id), str(request.build_absolute_uri())], stdin=subprocess.PIPE)

    messages.success(request, 'Progetto inviato ad agenzia')
    return redirect('flyer_builder_manager:projects')


@login_required
def catalog_products(request):
    code = request.GET.get('code', None)
    description = request.GET.get('description', None)
    seller_id = request.user.profile.client.seller_code
    prods = Product.objects
    seller=""
    if 'seller' in request.GET:
        seller=request.GET['seller']

    if seller != '0' and seller != "":
        prods = prods.filter(product_id__in=Distribution.objects.filter(seller=seller).values('product_id'))

    if (code is not None) and (code != ''):
        prods = prods.filter(product_id__in=Distribution.objects.filter(
            code__icontains=code).values('product_id'))
    if (description is not None) and (description != ''):
        prods = prods.filter(product_id__in=Description.objects.filter(
            Q(description__icontains=description)).values('product_id'))

    if not request.user.is_superuser:
        seller_id = request.user.profile.client.seller_code
        prods = prods.filter(product_id__in=Distribution.objects.filter(
            seller__id=seller_id).values('product_id'))

    if request.user.is_superuser and (not seller):
        product_list = Product.objects.all().order_by('-lastModify')
    else:
        product_list = prods.order_by('-lastModify')
    paginator = Paginator(product_list, 20)

    page_number = request.GET.get('page')

    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    sellers=Seller.objects.all()
    context = {'page_obj': page_obj,'sellers':sellers}
    return render(request, 'flyer_builder_manager/catalog/products.html', context)


@login_required
def delete_catalog_product_image(request, product_id, image_id):
    seller_id = request.user.profile.client.seller_code
    seller = Seller.objects.get(pk=seller_id)
    product_objects = Product.objects
    product_objects.filter(product_id__in=Distribution.objects.filter(
        seller__id=seller_id).values('product_id'))

    product = product_objects.get(pk=product_id)
    product_distributions = Distribution.objects.filter(
        product=product, seller=seller)

    picture = product.picture.get(pk=image_id)
    keep_picture = False
    picture_products = picture.product_set.all()
    for picture_product in picture_products:
        if keep_picture:
            break
        picture_product_distributions = Distribution.objects.filter(
            product=picture_product)
        for picture_product_distribution in picture_product_distributions:
            if picture_product_distribution.seller != seller:
                keep_picture = True
                break
    if not keep_picture:
        picture.delete()
        messages.success(request, 'Immagine eliminata!')
    else:
        messages.error(request, 'Errore: impossibile eiminare l\'immagine')

    return redirect('flyer_builder_manager:edit_product', product_id=product.pk)


@login_required
def delete_catalog_product(request, product_id):
    seller_id = request.user.profile.client.seller_code
    seller = Seller.objects.get(pk=seller_id)
    product_objects = Product.objects
    product_objects.filter(product_id__in=Distribution.objects.filter(
        seller__id=seller_id).values('product_id'))

    product = product_objects.get(pk=product_id)

    product_distributions = Distribution.objects.filter(
        product=product, seller=seller)
    for product_distribution in product_distributions:
        product_distribution.delete()

    descriptions = Description.objects.filter(product=product)
    for description in descriptions:
        description.delete()

    pictures = product.picture.all()
    for picture in pictures:
        keep_picture = False
        picture_products = picture.product_set.all()
        for picture_product in picture_products:
            if keep_picture:
                break
            picture_product_distributions = Distribution.objects.filter(
                product=picture_product)
            for picture_product_distribution in picture_product_distributions:
                if picture_product_distribution.seller != seller:
                    keep_picture = True
                    break
        if not keep_picture:
            picture.delete()

    product.delete()

    messages.success(request, 'Prodotto eliminato!')
    return redirect('flyer_builder_manager:catalog_products')


@login_required
def edit_product(request, product_id):
    # seller_id = request.user.profile.seller_code
    seller_id = request.user.profile.client.seller_code
    product_objects = Product.objects
    product_objects.filter(product_id__in=Distribution.objects.filter(
        seller__id=seller_id).values('product_id'))
    product = product_objects.get(pk=product_id)
    categories = Category.objects.all().order_by('name')
    subcategories = Subcategory.objects.filter(
        category=product.category).order_by('name')
    seller = Seller.objects.get(pk=seller_id)
    distribution = Distribution.objects.filter(
        product=product, seller=seller)[0]
    if Description.objects.filter(product=product, descriptionFieldNum=0).exists():
        field1 = Description.objects.filter(
            product=product, descriptionFieldNum=0)[0].description
    else:
        field1 = ""
    if Description.objects.filter(product=product, descriptionFieldNum=1).exists():
        field2 = Description.objects.filter(
            product=product, descriptionFieldNum=1)[0].description
    else:
        field2 = ""
    if Description.objects.filter(product=product, descriptionFieldNum=2).exists():
        field3 = Description.objects.filter(
            product=product, descriptionFieldNum=2)[0].description
    else:
        field3 = ""
    if Description.objects.filter(product=product, descriptionFieldNum=3).exists():
        field4 = Description.objects.filter(
            product=product, descriptionFieldNum=3)[0].description
    else:
        field4 = ""
    if request.POST:
        code = request.POST["code"]
        if code is not None:
            distribution.code = code
            distribution.save()
        field1 = request.POST["field1"]
        if field1 is not None:
            if Description.objects.filter(product=product, descriptionFieldNum=0).exists():
                desc = Description.objects.filter(
                    product=product, descriptionFieldNum=0)[0]
                desc.description = field1
                desc.save()
            else:
                desc = Description(
                    product=product, description=field1, descriptionFieldNum=0)
                desc.save()
        field2 = request.POST["field2"]
        if field2 is not None:
            if Description.objects.filter(product=product, descriptionFieldNum=1).exists():
                desc = Description.objects.filter(
                    product=product, descriptionFieldNum=1)[0]
                desc.description = field2
                desc.save()
            else:
                desc = Description(
                    product=product, description=field2, descriptionFieldNum=1)
                desc.save()
        field3 = request.POST["field3"]
        if field3 is not None:
            if Description.objects.filter(product=product, descriptionFieldNum=2).exists():
                desc = Description.objects.filter(
                    product=product, descriptionFieldNum=2)[0]
                desc.description = field3
                desc.save()
            else:
                desc = Description(
                    product=product, description=field3, descriptionFieldNum=2)
                desc.save()
        field4 = request.POST["field4"]
        if field4 is not None:
            if Description.objects.filter(product=product, descriptionFieldNum=3).exists():
                desc = Description.objects.filter(
                    product=product, descriptionFieldNum=3)[0]
                desc.description = field4
                desc.save()
            else:
                desc = Description(
                    product=product, description=field4, descriptionFieldNum=3)
                desc.save()
        category = request.POST["category"]
        if category is not None:
            product.category = Category.objects.get(pk=category)
            product.save()
        subcategory = request.POST["subcategory"]
        if subcategory is not None:
            product.subcategory = Subcategory.objects.get(pk=subcategory)
            product.save()
        gramm_type = request.POST["gramm_type"]
        if gramm_type is not None:
            if gramm_type == '':
                product.grammageType = None
            else:
                product.grammageType = gramm_type
            product.save()
        gramm_value = request.POST["gramm_value"]
        if gramm_value is not None:
            if gramm_value == '':
                product.grammageValue = None
            else:
                product.grammageValue = int(gramm_value)
            product.save()
        if 'branded_product' in request.POST:
            branded_product = bool(request.POST['branded_product'])
        else:
            branded_product = False
        product.branded_product = branded_product
        product.save()

        # favorite_picture = request.POST["favorite_picture"]
        # if favorite_picture is not None:
        #     pictures = product.picture.all()
        #     for picture in pictures:
        #         if picture.pk == int(favorite_picture):
        #             picture.favorite = True
        #         else:
        #             picture.favorite = False
        #         picture.save()

        if len(request.FILES) != 0:
            api_url = "http://31.7.149.101:9051/api/add_image"
            uploaded_file = request.FILES["image"]
            file_to_send = {'image': uploaded_file}
            response = requests.post(api_url, files=file_to_send)
            picture_path = response.text
            if not picture_path:
                messages.error(request, 'Errore: formato immagine non valido')
                return redirect('flyer_builder_manager:edit_product', product_id=product.pk)
            # picture = product.picture.filter(favorite=True)[0]
            # picture.path = picture_path
            # picture.save()
            picture = Picture(path=picture_path, favorite=True)
            picture.save()
            product.picture.add(picture)
            product.save()

        messages.success(request, "Prodotto modificato")
        return redirect('flyer_builder_manager:edit_product', product_id=product.pk)
    else:
        context = {'subcategories': subcategories, 'field1': field1, 'field2': field2, 'field3': field3, 'field4': field4, 'categories': categories,
                   'product': product, 'distribution': distribution}
        
        return render(request, 'flyer_builder_manager/product/edit_product.html', context)
            
@login_required
def show_product(request, product_id):
    # seller_id = request.user.profile.seller_code
    seller_id = request.user.profile.client.seller_code
    product_objects = Product.objects
    product_objects.filter(product_id__in=Distribution.objects.filter(
        seller__id=seller_id).values('product_id'))
    product = product_objects.get(pk=product_id)
    categories = Category.objects.all().order_by('name')
    subcategories = Subcategory.objects.filter(
        category=product.category).order_by('name')
    seller = Seller.objects.get(pk=seller_id)
    distribution = Distribution.objects.filter(
        product=product, seller=seller)[0]
    if Description.objects.filter(product=product, descriptionFieldNum=0).exists():
        field1 = Description.objects.filter(
            product=product, descriptionFieldNum=0)[0].description
    else:
        field1 = ""
    if Description.objects.filter(product=product, descriptionFieldNum=1).exists():
        field2 = Description.objects.filter(
            product=product, descriptionFieldNum=1)[0].description
    else:
        field2 = ""
    if Description.objects.filter(product=product, descriptionFieldNum=2).exists():
        field3 = Description.objects.filter(
            product=product, descriptionFieldNum=2)[0].description
    else:
        field3 = ""
    if Description.objects.filter(product=product, descriptionFieldNum=3).exists():
        field4 = Description.objects.filter(
            product=product, descriptionFieldNum=3)[0].description
    else:
        field4 = ""
    if request.POST:
        code = request.POST["code"]
        if code is not None:
            distribution.code = code
            distribution.save()
        field1 = request.POST["field1"]
        if field1 is not None:
            if Description.objects.filter(product=product, descriptionFieldNum=0).exists():
                desc = Description.objects.filter(
                    product=product, descriptionFieldNum=0)[0]
                desc.description = field1
                desc.save()
            else:
                desc = Description(
                    product=product, description=field1, descriptionFieldNum=0)
                desc.save()
        field2 = request.POST["field2"]
        if field2 is not None:
            if Description.objects.filter(product=product, descriptionFieldNum=1).exists():
                desc = Description.objects.filter(
                    product=product, descriptionFieldNum=1)[0]
                desc.description = field2
                desc.save()
            else:
                desc = Description(
                    product=product, description=field2, descriptionFieldNum=1)
                desc.save()
        field3 = request.POST["field3"]
        if field3 is not None:
            if Description.objects.filter(product=product, descriptionFieldNum=2).exists():
                desc = Description.objects.filter(
                    product=product, descriptionFieldNum=2)[0]
                desc.description = field3
                desc.save()
            else:
                desc = Description(
                    product=product, description=field3, descriptionFieldNum=2)
                desc.save()
        field4 = request.POST["field4"]
        if field4 is not None:
            if Description.objects.filter(product=product, descriptionFieldNum=3).exists():
                desc = Description.objects.filter(
                    product=product, descriptionFieldNum=3)[0]
                desc.description = field4
                desc.save()
            else:
                desc = Description(
                    product=product, description=field4, descriptionFieldNum=3)
                desc.save()
        category = request.POST["category"]
        if category is not None:
            product.category = Category.objects.get(pk=category)
            product.save()
        subcategory = request.POST["subcategory"]
        if subcategory is not None:
            product.subcategory = Subcategory.objects.get(pk=subcategory)
            product.save()
        gramm_type = request.POST["gramm_type"]
        if gramm_type is not None:
            if gramm_type == '':
                product.grammageType = None
            else:
                product.grammageType = gramm_type
            product.save()
        gramm_value = request.POST["gramm_value"]
        if gramm_value is not None:
            if gramm_value == '':
                product.grammageValue = None
            else:
                product.grammageValue = int(gramm_value)
            product.save()
        if 'branded_product' in request.POST:
            branded_product = bool(request.POST['branded_product'])
        else:
            branded_product = False
        product.branded_product = branded_product
        product.save()

        # favorite_picture = request.POST["favorite_picture"]
        # if favorite_picture is not None:
        #     pictures = product.picture.all()
        #     for picture in pictures:
        #         if picture.pk == int(favorite_picture):
        #             picture.favorite = True
        #         else:
        #             picture.favorite = False
        #         picture.save()

        if len(request.FILES) != 0:
            api_url = "http://31.7.149.101:9051/api/add_image"
            uploaded_file = request.FILES["image"]
            file_to_send = {'image': uploaded_file}
            response = requests.post(api_url, files=file_to_send)
            picture_path = response.text
            if not picture_path:
                messages.error(request, 'Errore: formato immagine non valido')
                return redirect('flyer_builder_manager:show_product', product_id=product.pk)
            # picture = product.picture.filter(favorite=True)[0]
            # picture.path = picture_path
            # picture.save()
            picture = Picture(path=picture_path, favorite=True)
            picture.save()
            product.picture.add(picture)
            product.save()

        messages.success(request, "Prodotto modificato")
        return redirect('flyer_builder_manager:show_product', product_id=product.pk)
    else:
        context = {'subcategories': subcategories, 'field1': field1, 'field2': field2, 'field3': field3, 'field4': field4, 'categories': categories,
                   'product': product, 'distribution': distribution}
        
        return render(request, 'flyer_builder_manager/product/show_product.html', context)


@login_required
def add_product(request):
    categories = Category.objects.all().order_by('name')
    if request.POST:
        seller_id = request.user.profile.client.seller_code
        seller = Seller.objects.get(pk=seller_id)

        code = request.POST['code']
        if (code != '') and (Distribution.objects.filter(seller=seller, code=code).exists()):
            messages.error(request, 'Codice prodotto già esistente')
            context = {'categories': categories}
            return render(request, 'flyer_builder_manager/add_product.html', context)

        api_url = "http://31.7.149.101:9051/api/add_image"
        uploaded_file = request.FILES["image"]
        file_to_send = {'image': uploaded_file}
        response = requests.post(api_url, files=file_to_send)
        picture_path = response.text
        if not picture_path:
            messages.error(request, 'Errore: formato immagine non valido')
            return redirect('flyer_builder_manager:add_product')

        field1 = request.POST['field1']
        field2 = request.POST['field2']
        field3 = request.POST['field3']
        field4 = request.POST['field4']
        category_id = request.POST['category']
        subcategory_id = request.POST['subcategory']
        gramm_type = request.POST['gramm_type']
        gramm_value = request.POST['gramm_value']
        if 'branded_product' in request.POST:
            branded_product = bool(request.POST['branded_product'])
        else:
            branded_product = False

        product = Product()
        category = Category.objects.get(pk=category_id)
        product.category = category
        subcategory = Subcategory.objects.get(pk=subcategory_id)
        product.subcategory = subcategory
        if int(gramm_type) != -1:
            product.grammageType = gramm_type
            product.grammageValue = gramm_value
        product.branded_product = branded_product
        product.save()

        if code != '':
            distribution = Distribution(
                product=product, seller=Seller.objects.get(pk=seller_id), code=code)
            distribution.save()

        if field1.replace(' ', '') != '':
            description = Description(
                product=product, description=field1, descriptionFieldNum=0)
            description.save()

        if field2.replace(' ', '') != '':
            description = Description(
                product=product, description=field2, descriptionFieldNum=1)
            description.save()

        if field3.replace(' ', '') != '':
            description = Description(
                product=product, description=field3, descriptionFieldNum=2)
            description.save()

        if field4.replace(' ', '') != '':
            description = Description(
                product=product, description=field4, descriptionFieldNum=3)
            description.save()

        # api_url = "http://31.7.149.101:9051/api/add_image"
        # uploaded_file = request.FILES["image"]
        # file_to_send = {'image': uploaded_file}
        # response = requests.post(api_url, files=file_to_send)
        # picture_path = response.text
        picture = Picture(path=picture_path)
        picture.save()
        product.picture.add(picture)
        product.save()

        # DOPO AVER CARICATO LE IMMAGINI SU A&B TRAMITE API
        # if picturePaths.count > 0:
        #     for picturePath in picturePaths:
        #         if picturePath!='':
        #             picture = Picture(path = picturePath)
        #             picture.save()
        #             product.picture.add(picture)
        #             product.save()

        messages.success(request, 'Prodotto creato')
        return redirect('flyer_builder_manager:catalog_products')
    else:
        context = {'categories': categories}
        return render(request, 'flyer_builder_manager/add_product.html', context)
@login_required
def campaign(request,seller_code):
    name=""
    sell_in=""
    sell_out=""
    seller=""
    if 'name' in request.GET:
        name=request.GET['name']
    if 'sellin' in request.GET:
        sell_in=request.GET['sellin']
    if 'sellout' in request.GET:
        sell_out=request.GET['sellout']
    if 'seller' in request.GET:
        seller=request.GET['seller']

    if seller_code == 0:
        projects= Project.objects.all()
    else:
        projects= Project.objects.filter(seller=seller_code)

    if name != "":
        projects= projects.filter(name__icontains=name)
    if sell_in != "":
        projects= projects.filter(sell_in__icontains=sell_in)
    if sell_out != "":
        projects= projects.filter(sell_out__icontains=sell_out)
    if seller != '0' and seller != '':
        projects= projects.filter(seller=seller)

    print(name)
    paginator = Paginator(projects, 20)
    
    page_number = request.GET.get('page')

    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    sellers= Seller.objects.all()
    context = {'projects': page_obj,'sellers':sellers}
    return render(request, 'flyer_builder_manager/campaign/list.html', context)

@login_required
def download(request, project_id, seller):
    # Prendo Tutti i tipi di file e il progetto e li mando al template download.html #####
    project = Project.objects.get(pk= project_id)
    seller = Seller.objects.get(pk= seller)
    files = FileUploaded.objects.filter(id_project = project_id)
    context = {'files' : files , 'project':project, 'seller':seller}
    return render(request,'flyer_builder_manager/campaign/download.html' , context)

@login_required
def search_header_banner(request):
    client = Client.objects.get(pk= request.GET['clientID'])
    banners = BannerImage.objects.filter(client= client)
    context = {
        'banners' : banners,
        'header_banner_id':request.GET['header_bannerID'],
        'page_id':request.GET['pageID'],
        'project_id':request.GET['projectID'],
        'type':request.GET['tipo']
    }
    return render(request, 'flyer_builder_manager/_search_image_result.html', context)

@login_required
def import_image_to_cell(request):
    special_cell_id= request.GET['specialCellID']
    imageID= request.GET['imageID']
    project_id=request.GET['project_id']
    page_id=request.GET['page_id']
    image = BannerImage.objects.get(pk = imageID)
    special_cell= ProjectPageSpecialCell.objects.get(pk = special_cell_id,type=request.GET['type'])
    special_cell.image = image
    special_cell.save()
    

    return JsonResponse({"status": "imported",'url':special_cell.image.image.url})


@login_required
def get_project_product_style_data(request, project_id, product_id):
    project = request.user.profile.client.projects.get(pk=project_id)
    product = project.products.get(pk=product_id)
    return JsonResponse({
            "price_color": product.price_color(),
            "description1_color": product.description_brand_color(),
            "description2_color": product.description_type_color(),
            "description3_color": product.description_tastes_color(),
            "description4_color": product.description_weight_color(),
            "price_style": product.price_style(),
            "description1_style": product.description_brand_style(),
            "description2_style": product.description_type_style(),
            "description3_style": product.description_tastes_style(),
            "description4_style": product.description_weight_style(),
            "price_integer_font": product.price_integer_font(),
            "price_float_font": product.price_float_font(),
            "description1_font": product.description1_font(),
            "description_font": product.description2_font(),
            "price_integer_font_size": product.price_integer_font_size(),
            "price_float_font_size": product.price_float_font_size(),
            "description1_font_size": product.description1_font_size(),
            "description_font_size": product.description2_font_size(),
    })


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def edit_project_product_style_data(request, project_id, product_id):
    project = request.user.profile.client.projects.get(pk=project_id)
    product = project.products.get(pk=product_id)
    price_color = request.POST["price_color"]
    description_brand_color = request.POST["description1_color"]
    description_type_color = request.POST["description2_color"]
    description_tastes_color = request.POST["description3_color"]
    description_weight_color = request.POST["description4_color"]
    price_style = request.POST["price_style"]
    description_brand_style = request.POST["description1_style"]
    description_type_style = request.POST["description2_style"]
    description_tastes_style = request.POST["description3_style"]
    description_weight_style = request.POST["description4_style"]
    price_int_font = request.POST["price_int_font"]
    price_int_font_size = request.POST["price_int_font_size"]
    price_float_font = request.POST["price_float_font"]
    price_float_font_size = request.POST["price_float_font_size"]
    description1_font = request.POST["description1_font"]
    description1_font_size = request.POST['description1_font_size']
    description_font = request.POST["description_font"]
    description_font_size = request.POST['description_font_size']
    apply_for = request.POST['apply_for']
    if apply_for == "applyProject":
        if not project.product_style:
            product_style = ProjectProductStyle.objects.create()
            project.product_style = product_style
            project.save()
        product_style = project.product_style
        product_style.price_color = price_color
        product_style.description_brand_color = description_brand_color
        product_style.description_type_color = description_type_color
        product_style.description_tastes_color = description_tastes_color
        product_style.description_weight_color = description_weight_color
        product_style.price_style = price_style
        product_style.description_brand_style = description_brand_style
        product_style.description_type_style = description_type_style
        product_style.description_tastes_style = description_tastes_style
        product_style.description_weight_style = description_weight_style
        product_style.price_integer_font = price_int_font
        product_style.price_integer_font_size = price_int_font_size
        product_style.price_float_font = price_float_font
        product_style.price_float_font_size = price_float_font_size
        product_style.description1_font = description1_font
        product_style.description1_font_size = description1_font_size
        product_style.description2_font = description_font
        product_style.description2_font_size = description_font_size
        product_style.save()
    if apply_for == "applyPage":
        if not product.cell.page.product_style:
            product_style = ProjectProductStyle.objects.create()
            page = product.cell.page
            page.product_style = product_style
            page.save()
        product_style = product.cell.page.product_style
        product_style.price_color = price_color
        product_style.description_brand_color = description_brand_color
        product_style.description_type_color = description_type_color
        product_style.description_tastes_color = description_tastes_color
        product_style.description_weight_color = description_weight_color
        product_style.price_style = price_style
        product_style.description_brand_style = description_brand_style
        product_style.description_type_style = description_type_style
        product_style.description_tastes_style = description_tastes_style
        product_style.description_weight_style = description_weight_style
        product_style.price_integer_font = price_int_font
        product_style.price_integer_font_size = price_int_font_size
        product_style.price_float_font = price_float_font
        product_style.price_float_font_size = price_float_font_size
        product_style.description1_font = description1_font
        product_style.description1_font_size = description1_font_size
        product_style.description2_font = description_font
        product_style.description2_font_size = description_font_size
        product_style.save()

    if not product.product_style:
        product_style = ProjectProductStyle.objects.create()
        product.product_style = product_style
        product.save()
    product_style = product.product_style
    product_style.price_color = price_color
    product_style.description_brand_color = description_brand_color
    product_style.description_type_color = description_type_color
    product_style.description_tastes_color = description_tastes_color
    product_style.description_weight_color = description_weight_color
    product_style.price_style = price_style
    product_style.description_brand_style = description_brand_style
    product_style.description_type_style = description_type_style
    product_style.description_tastes_style = description_tastes_style
    product_style.description_weight_style = description_weight_style
    product_style.price_integer_font = price_int_font
    product_style.price_integer_font_size = price_int_font_size
    product_style.price_float_font = price_float_font
    product_style.price_float_font_size = price_float_font_size
    product_style.description1_font = description1_font
    product_style.description1_font_size = description1_font_size
    product_style.description2_font = description_font
    product_style.description2_font_size = description_font_size
    product_style.save()

    product.project.update_last_modification_time()
    return JsonResponse({
        "status": "edited",
    })


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def edit_project_cell_style_data(request, project_id, page_id):
    project = request.user.profile.client.projects.get(pk=project_id)
    page = project.pages.get(number=page_id)
    border_width = request.POST["border_width"]
    border_color = request.POST["border_color"]
    border_style = request.POST["border_style"]
    apply_for = request.POST['apply_for']
    if apply_for == "applyProject":
        if not project.product_page_style:
            page_style = ProjectPageStyle.objects.create()
            project.product_page_style = page_style
            project.save()
        page_style = project.product_page_style
        page_style.border_width = border_width
        page_style.border_color = border_color
        page_style.border_style = border_style
        page_style.save()

    if not page.product_page_style:
        page_style = ProjectPageStyle.objects.create()
        page.product_page_style = page_style
        page.save()
    page_style = page.product_page_style
    page_style.border_width = border_width
    page_style.border_color = border_color
    page_style.border_style = border_style
    page_style.save()

    project.update_last_modification_time()
    return JsonResponse({
        "status": "edited",
    })


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def edit_project_page_style_data(request, project_id, page_id):
    project = request.user.profile.client.projects.get(pk=project_id)
    page = project.pages.get(number=page_id)
    header_per = int(request.POST["header_per"])
    body_per = int(request.POST["body_per"])
    footer_per = int(request.POST["footer_per"])
    apply_for = request.POST['apply_for']
    if apply_for == "applyProject":
        if not project.product_page_style:
            page_style = ProjectPageStyle.objects.create()
            project.product_page_style = page_style
            project.save()
        page_style = project.product_page_style
        page_style.header_per = header_per
        page_style.body_per = body_per
        page_style.footer_per = footer_per
        page_style.save()

    if not page.product_page_style:
        page_style = ProjectPageStyle.objects.create()
        page.product_page_style = page_style
        page.save()
    page_style = page.product_page_style
    page_style.header_per = header_per
    page_style.body_per = body_per
    page_style.footer_per = footer_per
    page_style.save()
    cells = page.cells.all()
    products_to_migrate = []
    for cell in cells:
        if cell.has_product():
            products_to_migrate.append(cell.product())
        cell.delete()
    page.create_cells()
    new_cells = page.cells.order_by('y_top_left_coord', 'x_top_left_coord')
    for idx, cell in enumerate(new_cells):
        if idx < len(products_to_migrate):
            products_to_migrate[idx].cell = cell
            products_to_migrate[idx].save()

    project.update_last_modification_time()
    return JsonResponse({
        "status": "edited",
    })

@login_required
@require_http_methods(["POST"])
@csrf_exempt
def update_page_image(request, project_id, page_id):
    project = request.user.profile.client.projects.get(pk=project_id)
    page = project.pages.get(number=page_id)
    image_type = request.POST['type']

    if request.FILES:
        file = request.FILES['file']
        image = SpecialImage(type=image_type, client=request.user.profile.client, category=project.category)
        image.image = file
        image.save()
    else:
        image_id = request.POST['image_id']
        image = SpecialImage(id=image_id)

    banner = ProjectPageSpecialCell.objects.get(page=page, type=image_type)
    banner.image = image
    banner.save()

    project.update_last_modification_time()
    return JsonResponse({
        "status": "edited",
    })


def preview_project(request, project_id, *args, **kwargs):
    project = request.user.profile.client.projects.get(pk=project_id)
    pages = project.pages.all().order_by('number')
    return render(request, 'flyer_builder_manager/project/preview_project_pdf.html', {'project':project, 'pages':pages})


@login_required
def project_product_list(request, project_id):
    params = request.GET
    project = request.user.profile.client.projects.get(pk=project_id)
    products = project.products.all()
    if 'qf' in params and params['qf'] != '':
        products = products.filter(
            Q(code__icontains=params['qf'].strip()) |
            Q(description__icontains=params['qf'].strip()) |
            Q(cell__page__number__icontains=params['qf'].strip()) 
        )
    page = params['page'] if 'page' in params else 1
    products = custom_pagination(page, products, limit=5)
    context = {'project': project, 'products': products,'project_id':project_id}
    return render(request, 'flyer_builder_manager/project/product_list.html', context)


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def save_page(request, project_id, *args, **kwargs):
    files = eval(request.POST['photos'])
    pages = eval(request.POST['pages'])
    # try:
    #     os.mkdir(f"media/project")
    # except:
    #     pass
    # try:
    #     os.mkdir(f"media/project/{project_id}")
    # except:
    #     pass
    # try:
    #     os.mkdir(f"media/project/{project_id}/images/")
    # except:
    #     pass
    for i, file in enumerate(files):
        page = ProjectPage.objects.get(number=pages[i], project_id=project_id)
        file = file.replace('data:image/jpeg;base64,', '')
        data = ContentFile(base64.b64decode(file))
        file_name = f'{pages[i]}.jpeg'
        page.image.save(file_name, data, save=True)

        # with open(file_name, "wb") as fh:
        #     fh.write(base64.b64decode(file))
    return JsonResponse({
        "status": "edited",
    })


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def update_product(request, project_id, product_id):
    project = request.user.profile.client.projects.get(pk=project_id)
    product = project.products.get(pk=product_id)
    value = request.POST.get('value', False)
    activity = request.POST.get('activity')
    if activity == "stopper":
        if value == "true":
            product.stopper = True
            ProjectStopper.objects.create(project=project, product=product, number=1)
        else:
            product.stopper = False
            ProjectStopper.objects.filter(project=project, product=product).delete()
    if activity == "poster":
        if value == "true":
            product.poster = True
            ProjectPoster.objects.create(project=project, product=product, number=1)
        else:
            product.poster = False
            ProjectPoster.objects.filter(project=project, product=product).delete()

    product.save()
    project.update_last_modification_time()
    return JsonResponse({"status": "Successfully updated" })
        
        
@login_required
def delete_product(request, project_id,product_id):
    project = request.user.profile.client.projects.get(pk=project_id)
    product = project.products.get(pk=product_id)
    product.delete()
    messages.success(request, 'Prodotto eliminato')
    return redirect(f'/project/{project_id}/product_list')


@login_required
def project_page_pdf_partial(request, project_id):
    project = request.user.profile.client.projects.get(pk=project_id)
    context = {'project': project}
    return render(request, 'flyer_builder_manager/project/page/_page.html', context)


@login_required
def project_stopper(request, project_id):
    project = request.user.profile.client.projects.get(pk=project_id)
    stoppers = project.stoppers.all()
    category = project.category
    default_header_images = SpecialImage.objects.filter(type="stopper_header",category=category, client=None)
    default_footer_images = SpecialImage.objects.filter(type="stopper_footer",category=category, client=None)
    user_header_images = request.user.profile.client.specialimage_set.filter(type='stopper_header', category=category)
    user_footer_images = request.user.profile.client.specialimage_set.filter(type='stopper_footer', category=category)
    header_images=default_header_images |user_header_images
    footer_images = default_footer_images |user_footer_images
    
    context = {'project': project, 'stoppers': stoppers, 'header_images':header_images, 'footer_images':footer_images, "stopper_reload":True,}
    return render(request, 'flyer_builder_manager/project/stopper_product.html', context)

@login_required
def project_poster(request, project_id):
    project = request.user.profile.client.projects.get(pk=project_id)
    posters = project.posters.all()
    category = project.category
    default_header_images = SpecialImage.objects.filter(type="poster_header",category=category, client=None)
    default_footer_images = SpecialImage.objects.filter(type="poster_footer",category=category, client=None)
    user_header_images = request.user.profile.client.specialimage_set.filter(type='poster_header', category=category)
    user_footer_images = request.user.profile.client.specialimage_set.filter(type='poster_footer', category=category)
    header_images=default_header_images |user_header_images
    footer_images = default_footer_images |user_footer_images
    
    context = {'project': project, 'posters': posters, 'header_images':header_images, 'footer_images':footer_images,"poster_reload":True}
    return render(request, 'flyer_builder_manager/project/poster_product.html', context)


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def edit_project_style_data(request, project_id, element):
    project = request.user.profile.client.projects.get(pk=project_id)
    header_per = int(request.POST["header_per"])
    body_per = int(request.POST["body_per"])
    footer_per = int(request.POST["footer_per"])
    if element == "stopper":
        if not project.stopper_style:
            stopper_style = ProjectPageStyle.objects.create()
            project.stopper_style = stopper_style
            project.save()
        stopper_style = project.stopper_style
        stopper_style.header_per = header_per
        stopper_style.body_per = body_per
        stopper_style.footer_per = footer_per
        stopper_style.save()
    elif element == "poster":
        if not project.poster_style:
            poster_style = ProjectPageStyle.objects.create()
            project.poster_style = poster_style
            project.save()
        poster_style = project.poster_style
        poster_style.header_per = header_per
        poster_style.body_per = body_per
        poster_style.footer_per = footer_per
        poster_style.save()

    project.update_last_modification_time()
    return JsonResponse({
        "status": "edited",
    })

@login_required
@require_http_methods(["GET"])
def project_partial(request, project_id, element):
    project = request.user.profile.client.projects.get(pk=project_id)
    if element == "stopper":
        stoppers = project.stoppers.all()
        header_images = SpecialImage.objects.filter(type="stopper_header")
        footer_images = SpecialImage.objects.filter(type="stopper_footer")
        context = {'project': project, 'stoppers': stoppers, "header_images":header_images, "footer_images":footer_images}
        return render(request, 'flyer_builder_manager/project/_stopper_poster.html', context)
    elif element == "poster":
        posters = project.posters.all()
        header_images = SpecialImage.objects.filter(type="poster_header")
        footer_images = SpecialImage.objects.filter(type="poster_footer")
        context = {'project': project, 'posters': posters, "header_images":header_images, "footer_images":footer_images}
        return render(request, 'flyer_builder_manager/project/_stopper_poster.html', context)
    return JsonResponse({
        "status": "edited",
    })


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def update_element_image(request, project_id, element):
    project = request.user.profile.client.projects.get(pk=project_id)
    image_type = request.POST['type']
    if request.FILES:
        file = request.FILES['file']
        image = SpecialImage(type=image_type, client=request.user.profile.client)
        image.image = file
        image.save()
    else:
        image_id = request.POST['image_id']
        image = SpecialImage(id=image_id)
    if element == "stopper":
        stoppers = project.stoppers.all()
        if 'header' in image_type:
            stoppers.update(header_image=image)
        else:
            stoppers.update(footer_image=image)
    elif element=="poster":
        posters = project.posters.all()
        if 'header' in image_type:
            posters.update(header_image=image)
        else:
            posters.update(footer_image=image)

    project.update_last_modification_time()
    return JsonResponse({
        "status": "edited",
    })


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def save_stopper(request, project_id, *args, **kwargs):
    files = eval(request.POST['photos'])
    project = request.user.profile.client.projects.get(pk=project_id)
    stoppers = project.stoppers.all()
    for i, file in enumerate(files):
        stopper = stoppers[i]
        file = file.replace('data:image/jpeg;base64,', '')
        data = ContentFile(base64.b64decode(file))
        file_name = f'{i}.jpeg'
        stopper.image.save(file_name, data, save=True)

        # with open(file_name, "wb") as fh:
        #     fh.write(base64.b64decode(file))
    return JsonResponse({
        "status": "edited",
    })

@login_required
@require_http_methods(["POST"])
@csrf_exempt
def save_poster(request, project_id, *args, **kwargs):
    files = eval(request.POST['photos'])
    project = request.user.profile.client.projects.get(pk=project_id)
    posters = project.posters.all()
    for i, file in enumerate(files):
        poster = posters[i]
        file = file.replace('data:image/jpeg;base64,', '')
        data = ContentFile(base64.b64decode(file))
        file_name = f'{i}.jpeg'
        poster.image.save(file_name, data, save=True)

        # with open(file_name, "wb") as fh:
        #     fh.write(base64.b64decode(file))
    return JsonResponse({
        "status": "edited",
    })

@login_required
@require_http_methods(["POST"])
@csrf_exempt
def edit_data_position(request, project_id,element):
    project = request.user.profile.client.projects.get(pk=project_id)
    # import pdb;pdb.set_trace()
    obj_id = request.POST['obj_id']
    ele_type = request.POST['type']
    top = int(request.POST["top"])
    left = int(request.POST["left"])
    if element == "stopper":
        stopper = project.stoppers.get(pk=obj_id)
        if ele_type == "image":
            stopper.image_top = top
            stopper.image_left = left
        else:
            stopper.info_top = top
            stopper.info_left = left
        stopper.save()
    elif element == "poster":
        poster = project.posters.get(pk=obj_id)
        if ele_type == "image":
            poster.image_top = top
            poster.image_left = left
        else:
            poster.info_top = top
            poster.info_left = left
        poster.save()
        
    project.update_last_modification_time()
    return JsonResponse({
        "status": "edited",
    })
   